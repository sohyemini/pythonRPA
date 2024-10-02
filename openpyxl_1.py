import openpyxl

wb = openpyxl.Workbook()        # create workbook

ws = wb.active                  # create 1st worksheet
ws['A1'] = 'Hello OpenPyXl'
ws['B1'] = 10
ws['B2'] = 20
ws['b3'] = '=B1+b2'

ws2 = wb.create_sheet('Hello')  # create new worksheet
ws2.cell(5, 2, "My first openpyxl program")

filename = "hello.xlsx"
wb.save(filename)

print("mission completed !!!")



