import openpyxl

wb = openpyxl.load_workbook(r"./data/us-colleges-and-universities.xlsx")
ws = wb['Feuil1']

ws_max_column = ws.max_column
ws_max_row = ws.max_row

print("No", ws['E1'].value, ws['F1'].value, ws['G1'].value, ws['F1'].value,
      ws['G1'].value, ws['I1'].value)

for i in range(2, ws_max_row+1, 1):
    print(i-1, ws[f'E{i}'].value, ws[f'F{i}'].value,
          ws[f'G{i}'].value, ws[f'F{i}'].value,
          ws[f'G{i}'].value, ws[f'I{i}'].value)

print(f"Total row = {ws_max_row}")
print(f"Total column = {ws_max_column}")