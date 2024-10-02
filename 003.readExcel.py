import openpyxl

wb = openpyxl.load_workbook(r"./data/us-colleges-and-universities.xlsx")

ws = wb['Feuil1']

print(ws['e1'].value, ws['f1'].value)
print(ws['e2'].value, ws['f2'].value)