import openpyxl

wb = openpyxl.load_workbook(r"./data/us-colleges-and-universities.xlsx")
ws = wb['Feuil1']

ws_max_column = ws.max_column
ws_max_row = ws.max_row

print("No", ws['E1'].value, ws['F1'].value, ws['G1'].value, ws['F1'].value,
      ws['G1'].value, ws['I1'].value)

for i in range(ws_max_row):
    print(i)